import openpyxl
import os

"""
On import les modules openpyxl et os.
On va l'utiliser pour ouvrir les fichier xlsx et ainsi les utiliser.
Il va servir a donner les chemins d'accès pour aller aux différents fichiers que nous allons utiliser.
"""

# --- LES FONCTIONS ---

def lecture_du_fichier_coef_dico(fichier_coef):
    """
    FONCTION POUR LIRE LE FICHIER EXCEL DES COEFS.

    **Explication du code :**
    
    1. **Création d'un tableau vide** pour l'utiliser par la suite.
    2. **On va aller ouvrir le fichier de refs** (le fichier Excel qui contient les coef).
    3. **Python lit la première ligne** de mon fichier Excel de ref qui contient les coefs (les entêtes).
    4. **On va faire une boucle** pour lire les données qui sont dans le fichiers.
    5. **Le zip** c'est pour dire que 'Matière'(entetes) : 'Maths' (ligne) , 'Coef'(entetes) : 3(ligne).
    6. **Et on transforme ca en dictionnaire**.
    7. **Append** : c'est pour rajouter tout ca au tableau.
    8. **On rempli le tableau** avec nos informations que l'on vient de trouver.

    :param fichier_coef: Le chemin du fichier Excel.
    :return: Le tableau rempli avec les coefficients.
    """
    # Création d'un tableau vide pour l'utiliser par la suite
    tableau_coef = []
    
    # On va aller ouvrir le fichier de refs
    Fichier_coef = openpyxl.load_workbook(fichier_coef, data_only=True)

    Onglets_coef = Fichier_coef.sheetnames
    feuille_active = Fichier_coef[Onglets_coef[0]] 

    # python lit la première ligne de mon fichier Excel de ref qui contient les coefs
    entetes = [cell.value for cell in feuille_active[1]]
        
    # On va faire une boucle pour lire les données qui sont dans le fichiers
    for ligne in feuille_active.iter_rows(
        min_row=2, max_row=feuille_active.max_row,
        values_only=True):
        
        # le zip c'est pour dire que 'Matière'(entetes) : 'Maths' (ligne)
        # et on transforme ca en dictionnaire
        # append c'est pour rajouter tout ca au tableau
        tableau_coef.append(dict(zip(entetes, ligne)))
    
    # on rempli le tableau avec nos informations que l'on vient de trouver
    return tableau_coef


def lire_fichier_excel(fichier, Dossier_semestre):
    """
    FONCTION POUR LIRE LES FICHIERS EXCEL DES NOTES DES ETUDIANTS.

    **Explication du code :**

    1. **On créer un tableau vide** pour mettre nos valeurs dedans.
    2. **Fichier_Matière** : c'est pour dire que le nom de la matière c'est son nom (ex : Mathematiques.xlsx).
    3. **Dossier_semestre** : Meme chose pour le dossier.
    4. **Fichier_Excel** : c'est pour avoir le chemin complet des fichiers xlsx.
    5. **[1:4]** : c'est pour dire les collonnes B , C , D.
    6. **On va lire les données** des fichiers xlsx.
    7. **On veut la ligne min 2** et on va lire la valeurs qui est dans la case [A2] qui correspond au nombre d'éléves.
    8. **On fait un dictionnaire** des valeurs des entetes et des lignes qui sont présent dans le fichier.
    9. **On remplie le tableau** avec les valeurs qu'on vient de trouver.

    :param fichier: Nom du fichier.
    :param Dossier_semestre: Dossier contenant le fichier.
    :return: Tableau des notes.
    """
    # on créer un tableau vide pour mettre nos valeurs dedans
    note_matiere = []

    # c'est pour dire que le nom de la matière c'est son nom
    Fichier_Matière = os.path.basename(fichier)
    # Meme chose pour le dossier
    Dossier_semestre_basename = os.path.basename(Dossier_semestre)

    # c'est pour avoir le chemin complet des fichiers xlsx
    Fichier_Excel = openpyxl.load_workbook(os.path.join(Dossier_semestre, fichier), data_only = True)
    
    Onglets = Fichier_Excel.sheetnames
    feuille_active = Fichier_Excel[Onglets[0]]

    # [1:4] c'est pour dire les collonnes B , C , D
    entetes = [cell.value for cell in feuille_active[1]][1:4]
        
    # on va lire les données des ficheirs xlsx (nombre d'élèves en A2)
    nblignes= feuille_active['A2'].value
    
    # on va lire les données dans le fichierx xlsx
    for ligne in feuille_active.iter_rows(
        min_row=2, max_row=nblignes, min_col=2, max_col=4,
        values_only=True):
        
        # on fait un dictionnaire des valeursdes entetes et des lignes
        note = dict(zip(entetes, ligne)) | {
            "Fichier_Matière": Fichier_Matière,
            "Dossier_semestre": Dossier_semestre_basename      
        }
        # on remplie le tableau avec les valeurs qu'on vient de trouver 
        note_matiere.append(note)   
    return note_matiere


def determiner_etat_ue(note):
    """
    Détermine l'état d'une UE selon sa moyenne annuelle.
    
    * Si note >= 10 : VALIDÉ
    * Si note >= 8 : COMPENSABLE
    * Sinon : NON VALIDÉ
    """
    if note >= 10:
        return "VALIDÉ", "ue-validee"
    elif note >= 8:
        return "COMPENSABLE", "ue-compensee"
    else:
        return "NON VALIDÉ", "ue-echec"


def calculer_decision_passage(liste_moyennes_annuelles):
    """
    Règles de passage BUT sur les MOYENNES ANNUELLES :
    
    1. Aucune UE < 8 (Pas de 'NON VALIDÉ')
    2. Au moins 2 UE >= 10
    """
    if not liste_moyennes_annuelles:
        return "Incomplet", "decision-fail"
    if any(note < 8 for note in liste_moyennes_annuelles):
        return "REFUSÉ (UE < 8)", "decision-fail"
    
    nb_ue_validees = sum(1 for note in liste_moyennes_annuelles if note >= 10)
    
    if nb_ue_validees >= 2:
        return "ADMIS (Passage BUT2)", "decision-ok"
    else:
        return "REFUSÉ (Manque UE > 10)", "decision-fail"


# --- PROGRAMME PRINCIPAL (Protégé pour Sphinx) ---
if __name__ == "__main__":
    
    #CHEMIN DES FICHIERS EXCEL
    fichier_de_ref = '/home/etudiant/.ssh/SAE105_EXCEL_-_VALIDATION_UE/data/coefficients/Coef.xlsx'

    dossiers_notes = []
    dossiers_notes.append('/home/etudiant/.ssh/SAE105_EXCEL_-_VALIDATION_UE/data/notes/notes_S1')
    dossiers_notes.append('/home/etudiant/.ssh/SAE105_EXCEL_-_VALIDATION_UE/data/notes/notes_S2')

    #INITIALISATION DES TABLEAUX
    tableau_coef = []
    note_matiere = []
    Gros_Tableau_Notes = []
    notes = {}

    # Lecture du fichier des coefs  
    tableau_coef = lecture_du_fichier_coef_dico(fichier_de_ref)

    # Lecture des fichiers des notes des étudiants
    for dossier_notes in dossiers_notes:    
        for fichier in os.listdir(dossier_notes):
            Gros_Tableau_Notes = Gros_Tableau_Notes + lire_fichier_excel(fichier , dossier_notes)

    # liste des UE
    liste_ue = list({item["Unité_d_Enseignement"] for item in tableau_coef})
    liste_semestre = list({item["Semestre"] for item in tableau_coef})

    for ue in liste_ue:
        for matière in tableau_coef:
            matiere_ue = matière["Unité_d_Enseignement"]
            matiere_fichier = matière["Fichier"]
            matiere_coef = float(matière["Coefficient"])
            if  matiere_ue == ue :
                for eleve in Gros_Tableau_Notes:
                    cle = (eleve["Nom"], eleve["Prénom"],ue)
                    if eleve["Fichier_Matière"] == matière["Fichier"] :
                        if cle not in notes:
                            notes[cle] = 0
                        notes[cle] += float(eleve["Note"]) * matiere_coef / 100

    ues_racines = sorted(list({ue.split('.')[0] for ue in liste_ue}))
    etudiants_uniques = sorted(list(set((k[0], k[1]) for k in notes.keys())))

    # GÉNÉRATION DU HTML 
    html_content = """
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <title>BUT1-R&T</title>
        <style>
            body{ font-family: 'Segoe UI', sans-serif; padding: 20px; font-size: 13px; }
            h2 { text-align: center; color: #333; }
            table{ border-collapse: collapse; width: 100%; box-shadow: 0 0 15px rgba(0,0,0,0.1); margin-top: 20px; }
            td, th{ border: 1px solid #ddd; padding: 8px 4px; text-align: center; }
            thead tr:first-child th { background-color: #005f99; color: white; border-right: 1px solid white; }
            thead tr:nth-child(2) th { background-color: #007acc; color: white; font-size: 0.9em; }
            .col-semestre { background-color: #fcfcfc; color: #555; }
            .col-moyenne { background-color: #fff; font-weight: bold; border-left: 2px solid #ccc; }
            .ue-validee { background-color: #d4edda; color: #155724; font-weight: bold; }
            .ue-compensee { background-color: #fff3cd; color: #856404; font-weight: bold; }
            .ue-echec { background-color: #f8d7da; color: #721c24; font-weight: bold; }
            .decision-ok { background-color: #28a745; color: white; font-weight: bold; font-size: 1.1em; }
            .decision-fail { background-color: #dc3545; color: white; font-weight: bold; font-size: 1.1em; }
            tr:hover { background-color: #f1f1f1; }
        </style>
    </head>
    <body>
        <h2>Notes BUT1 - R&T </h2>
        <table>
        <thead>
            <tr>
                <th rowspan="2" style="width: 150px;">Étudiant</th>
    """

    for ue in ues_racines:
        html_content += f'<th colspan="4">{ue} (Annuel)</th>'

    html_content += '<th rowspan="2">DÉCISION</th></tr>'
    html_content += '<tr>'
    for ue in ues_racines:
        html_content += f'<th>{ue}.1</th><th>{ue}.2</th><th>Moy</th><th>État</th>'
    html_content += '</tr></thead><tbody>'

    for (nom, prenom) in etudiants_uniques:
        moyennes_annuelles_etudiant = []
        ligne_html_etudiant = f"<tr><td style='text-align:left; padding-left:10px;'><b>{nom}</b> {prenom}</td>"
        
        for racine in ues_racines:
            nom_ue_s1 = f"{racine}.1"
            nom_ue_s2 = f"{racine}.2"
            note_s1 = notes.get((nom, prenom, nom_ue_s1), 0.0)
            note_s2 = notes.get((nom, prenom, nom_ue_s2), 0.0)
            moyenne_annuelle = (note_s1 + note_s2) / 2
            moyennes_annuelles_etudiant.append(moyenne_annuelle)
            txt_etat, class_etat = determiner_etat_ue(moyenne_annuelle)
            
            ligne_html_etudiant += f'<td class="col-semestre">{round(note_s1, 2)}</td>'
            ligne_html_etudiant += f'<td class="col-semestre">{round(note_s2, 2)}</td>'
            ligne_html_etudiant += f'<td class="col-moyenne">{round(moyenne_annuelle, 2)}</td>'
            ligne_html_etudiant += f'<td class="{class_etat}">{txt_etat}</td>'

        txt_decision, class_decision = calculer_decision_passage(moyennes_annuelles_etudiant)
        ligne_html_etudiant += f'<td class="{class_decision}">{txt_decision}</td></tr>'
        html_content += ligne_html_etudiant

    html_content += """
        </tbody>
        </table>
    </body>
    </html>
    """

    with open("/home/etudiant/.ssh/SAE105_EXCEL_-_VALIDATION_UE/html/mes_notes.html", "w", encoding="utf-8") as file:
        file.write(html_content)

    print("Le fichier 'mes_notes.html' a été généré avec succès.")