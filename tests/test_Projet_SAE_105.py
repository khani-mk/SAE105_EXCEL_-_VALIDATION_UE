#On import les modules openpyxl et os

#On va l'utiliser pour ouvrir les fichier xlsx et ainsi les utiliser
import openpyxl

#il va servir a donner les chemins d'accès pour aller aux différents fichiers que nous allons utiliser
import os


#CHEMIN DES FICHIERS EXCEL
#Chemin pour accéder au fichier excel des coefs
fichier_de_ref = '/workspaces/SAE105_EXCEL_-_VALIDATION_UE/tests/test/Coef.xlsx'

#CHEMIN POUR ACCEDER AU DOSSIER QUI CONTIENT LES FICHIERS EXCEL DES NOTES DE TOUS LES ETUDIANTS
 

dossiers_notes = []

dossiers_notes.append('/workspaces/SAE105_EXCEL_-_VALIDATION_UE/tests/test/notes_S1')
dossiers_notes.append('/workspaces/SAE105_EXCEL_-_VALIDATION_UE/tests/test/notes_S2')

#INITIALISATION DES TABLEAUX
# on créait des tableaux vide pour pouvoir mettre les données que nous voulons utilisées

#On créait un tableau_coef pour y mettre les données qui sont présent dans le fichiers Coef.xlsx pour que ca soit plus simple a utiliser
tableau_coef = []

#On créait un tableau_note pour y mettres les données qui sont présent dans les fichiers xlsx qui sont dans le dossier notes_S1 pour pouvoir les utiliser
note_matiere = []

#On créait un Gros_tableau_Notes qui va contenir les notes de chaque matière de chaque élèves que nous allons utiliser
Gros_Tableau_Notes = []

#C'est le dictionnaire que nous allons utiliser plus tard 
notes = {}



#FONCTION POUR LIRE LE FICHIER EXCEL DES COEFS
def lecture_du_fichier_coef_dico(fichier_coef):
    tableau_coef = []
    Fichier_coef= openpyxl.load_workbook(fichier_de_ref, data_only = True)

    Onglets_coef = Fichier_coef.sheetnames
    feuille_active = Fichier_coef[Onglets_coef[0]] # la feuille active est le premier onglet du fichier
    
    entetes = [cell.value for cell in feuille_active[1]]
        

    for ligne in feuille_active.iter_rows(
        min_row=2, max_row=feuille_active.max_row,
        values_only=True):
        tableau_coef.append(dict(zip(entetes, ligne)))
    
    return tableau_coef

#   FONCTION POUR LIRE LES FICHIERS EXCEL DES NOTES DES ETUDIANTS   
def lire_fichier_excel(fichier , Dossier_semestre):
    note_matiere = []
    Fichier_Matière = os.path.basename(fichier)
    Dossier_semestre_basename = os.path.basename(Dossier_semestre)
    Fichier_Excel = openpyxl.load_workbook(os.path.join(Dossier_semestre, fichier), data_only = True)
    Onglets = Fichier_Excel.sheetnames
    feuille_active = Fichier_Excel[Onglets[0]]
    entetes = [cell.value for cell in feuille_active[1]][1:4]

        
    nblignes= feuille_active['A2'].value
    for ligne in feuille_active.iter_rows(
        min_row=2, max_row=nblignes, min_col=2, max_col=4,
        values_only=True):
        note = dict(zip(entetes, ligne)) | {
            "Fichier_Matière": Fichier_Matière,
            "Dossier_semestre": Dossier_semestre_basename       
        }
        note_matiere.append(note)   
    return note_matiere
   
#Début du programme principal
# Lecture du fichier des coefs  
tableau_coef = lecture_du_fichier_coef_dico(fichier_de_ref)

# Lecture des fichiers des notes des étudiants
for dossier_notes in dossiers_notes:    
    for fichier in os.listdir(dossier_notes):
        Gros_Tableau_Notes = Gros_Tableau_Notes + lire_fichier_excel(fichier , dossier_notes)


# liste des UE


liste_ue = list({item["Unité_d_Enseignement"] for item in tableau_coef})

liste_semestre = list({item["Semestre"] for item in tableau_coef})



for ue in liste_ue:
    #print("================================== " , ue)
    for matière in tableau_coef:

        matiere_ue = matière["Unité_d_Enseignement"]
        matiere_fichier = matière["Fichier"]
        matiere_coef = float(matière["Coefficient"])
        #SI la matière existe dans l'UE en cours
        if  matiere_ue == ue :
            for eleve in Gros_Tableau_Notes:
                cle = (eleve["Nom"], eleve["Prénom"],ue)  # identifiant unique
                # on ne traite que la matière en cours 
                #print('>>>>>>',eleve["Fichier_Matière"])
                if eleve["Fichier_Matière"] == matière["Fichier"] :
                    #print("***",eleve["Fichier_Matière"],eleve["Note"],matiere_coef)
                    # Creation du tableau de notes final avec la pondération
                    if cle not in notes:
                        notes[cle] = 0
                    notes[cle] += float(eleve["Note"]) * matiere_coef / 100

    

# On détermine les noms des UEs "racines" (ex: UE1, UE2)
# On suppose le format "UEy.x". On split sur le point et on garde la première partie.
ues_racines = sorted(list({ue.split('.')[0] for ue in liste_ue}))

# Liste de tous les étudiants (Nom, Prénom) uniques
etudiants_uniques = sorted(list(set((k[0], k[1]) for k in notes.keys())))


def determiner_etat_ue(note):
    """
    Détermine l'état d'une UE selon sa moyenne annuelle.
    """
    if note >= 10:
        return "VALIDÉ", "ue-validee"
    elif note >= 8:
        return "COMPENSABLE", "ue-compensee"
    else:
        return "NON VALIDÉ", "ue-echec"

def calculer_decision_passage(liste_moyennes_annuelles):
    """
    Règles de passage BUT sur les MOYENNES ANNUELLES :
    1. Aucune UE < 8 (Pas de 'NON VALIDÉ')
    2. Au moins 2 UE >= 10
    """
    # S'il n'y a aucune note
    if not liste_moyennes_annuelles:
        return "Incomplet", "decision-fail"

    # 1. Vérification éliminatoire (< 8)
    if any(note < 8 for note in liste_moyennes_annuelles):
        return "REFUSÉ (UE < 8)", "decision-fail"
    
    # 2. Comptage des UE validées (>= 10)
    nb_ue_validees = sum(1 for note in liste_moyennes_annuelles if note >= 10)
    
    # Règle : Il faut au moins 2 UE >= 10
    if nb_ue_validees >= 2:
        return "ADMIS (Passage BUT2)", "decision-ok"
    else:
        return "REFUSÉ (Manque UE > 10)", "decision-fail"


# GÉNÉRATION DU HTML 

html_content = """
<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>BUT1-R&T</title>
    <style>
        body{ font-family: 'Segoe UI', sans-serif; padding: 20px; font-size: 13px; }
        h2 { text-align: center; color: #333; }
        table{ border-collapse: collapse; width: 100%; box-shadow: 0 0 15px rgba(0,0,0,0.1); margin-top: 20px; }
        
        /* Cellules */
        td, th{ border: 1px solid #ddd; padding: 8px 4px; text-align: center; }
        
        /* En-têtes */
        thead tr:first-child th { background-color: #005f99; color: white; border-right: 1px solid white; }
        thead tr:nth-child(2) th { background-color: #007acc; color: white; font-size: 0.9em; }
        
        /* Distinction visuelle des semestres */
        .col-semestre { background-color: #fcfcfc; color: #555; }
        .col-moyenne { background-color: #fff; font-weight: bold; border-left: 2px solid #ccc; }
        
        /* États UE */
        .ue-validee { background-color: #d4edda; color: #155724; font-weight: bold; }
        .ue-compensee { background-color: #fff3cd; color: #856404; font-weight: bold; }
        .ue-echec { background-color: #f8d7da; color: #721c24; font-weight: bold; }

        /* Décision finale */
        .decision-ok { background-color: #28a745; color: white; font-weight: bold; font-size: 1.1em; }
        .decision-fail { background-color: #dc3545; color: white; font-weight: bold; font-size: 1.1em; }
        
        tr:hover { background-color: #f1f1f1; }
    </style>
</head>
<body>
    <h2>Notes BUT1 - R&T </h2>
    
    <table>
    <thead>
        <tr>
            <th rowspan="2" style="width: 150px;">Étudiant</th>
"""

# Pour chaque UE Racine, on crée un titre qui s'étend sur 4 colonnes
for ue in ues_racines:
    html_content += f'<th colspan="4">{ue} (Annuel)</th>'

html_content += '<th rowspan="2">DÉCISION</th></tr>'

# -- LIGNE 2 : Les sous-colonnes (S1, S2, Moy, État) --
html_content += '<tr>'
for ue in ues_racines:
    # On suppose que les sous-UEs s'appellent UE1.1 et UE1.2
    html_content += f'<th>{ue}.1</th><th>{ue}.2</th><th>Moy</th><th>État</th>'
html_content += '</tr></thead><tbody>'

# --- REMPLISSAGE DU TABLEAU ---
for (nom, prenom) in etudiants_uniques:
    
    # On prépare une liste pour stocker les moyennes annuelles de cet étudiant
    # Cela servira à calculer la décision finale (Passage ou non)
    moyennes_annuelles_etudiant = []
    
    # Début de la ligne HTML pour l'étudiant
    ligne_html_etudiant = f"<tr><td style='text-align:left; padding-left:10px;'><b>{nom}</b> {prenom}</td>"
    
    # Pour chaque UE Racine (UE1, UE2...)
    for racine in ues_racines:
        nom_ue_s1 = f"{racine}.1"
        nom_ue_s2 = f"{racine}.2"
        
        # Récupération des notes (0.0 par défaut si manquant)
        note_s1 = notes.get((nom, prenom, nom_ue_s1), 0.0)
        note_s2 = notes.get((nom, prenom, nom_ue_s2), 0.0)
        
        # Calcul Moyenne Annuelle
        moyenne_annuelle = (note_s1 + note_s2) / 2
        
        # On ajoute à la liste pour la décision finale
        moyennes_annuelles_etudiant.append(moyenne_annuelle)
        
        # Détermination de l'état (Validé/Compensé/Echec)
        txt_etat, class_etat = determiner_etat_ue(moyenne_annuelle)
        
        # -- Ajout des 4 cellules dans le HTML --
        # 1. Note Semestre 1
        ligne_html_etudiant += f'<td class="col-semestre">{round(note_s1, 2)}</td>'
        # 2. Note Semestre 2
        ligne_html_etudiant += f'<td class="col-semestre">{round(note_s2, 2)}</td>'
        # 3. Moyenne Annuelle (Mise en valeur)
        ligne_html_etudiant += f'<td class="col-moyenne">{round(moyenne_annuelle, 2)}</td>'
        # 4. État
        ligne_html_etudiant += f'<td class="{class_etat}">{txt_etat}</td>'

    txt_decision, class_decision = calculer_decision_passage(moyennes_annuelles_etudiant)
    ligne_html_etudiant += f'<td class="{class_decision}">{txt_decision}</td></tr>'
    
    html_content += ligne_html_etudiant

html_content += """
    </tbody>
    </table>
</body>
</html>
"""

# Écriture dans le fichier
with open("mes_notes.html", "w", encoding="utf-8") as file:
    file.write(html_content)

print("Le fichier 'mes_notes.html' a été généré avec succès.")