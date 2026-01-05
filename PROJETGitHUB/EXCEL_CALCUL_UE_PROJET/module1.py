import openpyxl
import os

"""
On import les modules openpyxl et os.
On va l'utiliser pour ouvrir les fichier xlsx et ainsi les utiliser.
Il va servir a donner les chemins d'accès pour aller aux différents fichiers que nous allons utiliser.
"""

# --- LES FONCTIONS ---

def lecture_du_fichier_coef_dico(fichier_coef):
    """
    FONCTION POUR LIRE LE FICHIER EXCEL DES COEFS.

    **Explication du code :**
    
    1. **Création d'un tableau vide** pour l'utiliser par la suite.
    2. **On va aller ouvrir le fichier de refs** (le fichier Excel qui contient les coef).
    3. **Python lit la première ligne** de mon fichier Excel de ref qui contient les coefs (les entetes).
    4. **On va faire une boucle** pour lire les données qui sont dans le fichiers.
    5. **Le zip** c'est pour dire que 'Matière'(entetes) : 'Maths' (ligne) , 'Coef'(entetes) : 3(ligne).
    6. **Et on transforme ca en dictionnaire**.
    7. **Append** : c'est pour rajouter tout ca au tableau.
    8. **On rempli le tableau** avec nos informations que l'on vient de trouver.

    :param fichier_coef: Le chemin du fichier Excel.
    :return: Le tableau rempli avec les coefficients.
    """
    # Création d'un tableau vide pour l'utiliser par la suite
    tableau_coef = []
    
    # On va aller ouvrir le fichier de refs
    Fichier_coef = openpyxl.load_workbook(fichier_coef, data_only=True)

    Onglets_coef = Fichier_coef.sheetnames
    feuille_active = Fichier_coef[Onglets_coef[0]] 

    # python lit la première ligne de mon fichier Excel de ref qui contient les coefs
    entetes = [cell.value for cell in feuille_active[1]]
        
    # On va faire une boucle pour lire les données qui sont dans le fichiers
    for ligne in feuille_active.iter_rows(
        min_row=2, max_row=feuille_active.max_row,
        values_only=True):
        
        # le zip c'est pour dire que 'Matière'(entetes) : 'Maths' (ligne)
        # et on transforme ca en dictionnaire
        # append c'est pour rajouter tout ca au tableau
        tableau_coef.append(dict(zip(entetes, ligne)))
    
    # on rempli le tableau avec nos informations que l'on vient de trouver
    return tableau_coef

def lire_fichier_excel(fichier, Dossier_semestre):
    """
    FONCTION POUR LIRE LES FICHIERS EXCEL DES NOTES DES ETUDIANTS.

    **Explication du code :**

    1. **On créer un tableau vide** pour mettre nos valeurs dedans.
    2. **Fichier_Matière** : c'est pour dire que le nom de la matière c'est son nom (ex : Mathematiques.xlsx).
    3. **Dossier_semestre** : Meme chose pour le dossier.
    4. **Fichier_Excel** : c'est pour avoir le chemin complet des fichiers xlsx.
    5. **[1:4]** : c'est pour dire les collonnes B , C , D.
    6. **On va lire les données** des fichiers xlsx.
    7. **On veut la ligne min 2** et on va lire la valeurs qui est dans la case [A2] qui correspond au nombre d'éléves.
    8. **On fait un dictionnaire** des valeurs des entetes et des lignes qui sont présent dans le fichier.
    9. **On remplie le tableau** avec les valeurs qu'on vient de trouver.

    :param fichier: Nom du fichier.
    :param Dossier_semestre: Dossier contenant le fichier.
    :return: Tableau des notes.
    """
    # on créer un tableau vide pour mettre nos valeurs dedans
    note_matiere = []

    # c'est pour dire que le nom de la matière c'est son nom
    Fichier_Matière = os.path.basename(fichier)
    # Meme chose pour le dossier
    Dossier_semestre_basename = os.path.basename(Dossier_semestre)

    # c'est pour avoir le chemin complet des fichiers xlsx
    Fichier_Excel = openpyxl.load_workbook(os.path.join(Dossier_semestre, fichier), data_only = True)
    
    Onglets = Fichier_Excel.sheetnames
    feuille_active = Fichier_Excel[Onglets[0]]

    # [1:4] c'est pour dire les collonnes B , C , D
    entetes = [cell.value for cell in feuille_active[1]][1:4]
        
    # on va lire les données des ficheirs xlsx (nombre d'élèves en A2)
    nblignes= feuille_active['A2'].value
    
    # on va lire les données dans le fichierx xlsx
    for ligne in feuille_active.iter_rows(
        min_row=2, max_row=nblignes, min_col=2, max_col=4,
        values_only=True):
        
        # on fait un dictionnaire des valeursdes entetes et des lignes
        note = dict(zip(entetes, ligne)) | {
            "Fichier_Matière": Fichier_Matière,
            "Dossier_semestre": Dossier_semestre_basename      
        }
        # on remplie le tableau avec les valeurs qu'on vient de trouver 
        note_matiere.append(note)   
    return note_matiere


def calculer_decision_passage(liste_moyennes_annuelles):
    """
    Règles de passage BUT sur les MOYENNES ANNUELLES :
    
    1. Aucune UE < 8 (Pas de 'NON VALIDÉ')
    2. Au moins 2 UE >= 10
    """
    # S'il n'y a aucune note
    if not liste_moyennes_annuelles:
        return "Incomplet", "decision-fail"

    # On va vérifier si la note est inférieur a 8 si c'est le cas on va écrire REFUSE ( UE < 8)
    if any(note < 8 for note in liste_moyennes_annuelles):
        return "REFUSÉ (UE < 8)", "decision-fail"
    
    # On va vérifier si la notest supérieur ou égal a 10
    nb_ue_validees = sum(1 for note in liste_moyennes_annuelles if note >= 10)
    
    # Règle : Il faut au moins 2 UE >= 10
    #Si y a au minimum 2 UE qui sont supérieur a ou égal a 2 on écrit (" ADMIS ") si c'est pas le cas on écrit ("REFUSE")
    if nb_ue_validees >= 2:
        return "ADMIS (Passage BUT2)", "decision-ok"
    else:
        return "REFUSÉ (Manque UE > 10)", "decision-fail"


def determiner_etat_ue(note):
    """
    Détermine l'état d'une UE selon sa moyenne annuelle.
    
    * Si note >= 10 : VALIDÉ
    * Si note >= 8 : COMPENSABLE
    * Sinon : NON VALIDÉ
    """
    if note >= 10:
        return "VALIDÉ", "ue-validee"
    elif note >= 8:
        return "COMPENSABLE", "ue-compensee"
    else:
        return "NON VALIDÉ", "ue-echec"


def main():
    """
    PROGRAMME PRINCIPAL.

    **Ce que fait cette partie du code :**
    
    1. **CHEMIN DES FICHIERS EXCEL** : ``fichier_de_ref = '/workspaces/SAE105_EXCEL_-_VALIDATION_UE/data/coefficients/Coef.xlsx'``
    2. **On créait un tableau vide** pour mettre les données qu'on va vouloir utiliser (dossiers_notes).
    3. **Chemin pour accéder au note des S1 et S2**.
    4. **on met les données qui sont dans le fichier notes_S1 et notes_S2** dans le tableau que l'on vient de créer.
    5. **INITIALISATION DES TABLEAUX** : tableau_coef, note_matiere, Gros_Tableau_Notes, notes.
    6. **Lecture du fichier des coefs sous forme d'un dictionnaire**.
    7. **Lecture des fichiers des notes des étudiants**.
    8. **On va renvoier la liste des fichiers qui sont dans le tableau dossier_notes**.
    9. **dans le tableau Gros_Tableau_Notes on rajoute les données de fichier et de dossier_notes**.
    10. **liste des UE** : on va parcourir tableau_coef et pour chaque case ou y a écrit "Unité_d_Enseignement" on garde la valeur.
    11. **On fait la meme chose mais pour "Semestre"**.
    12. **on fait une boucle pour traiter tout les UE qui sont dans le liste_ue qui est dans le tableau_coef**.
    13. **on met en place des filtres**, **on fait une deuxième boucle**.
    14. **matiere_ue** : on va donner l'UE en fonction de la matière.
    15. **matiere_fichier** : on va donner la matière grace a la colonne fichier.
    16. **matiere_coef** : on va dire les coefs de la matière.
    17. **on filtre et on veut que ca soit uniquement les matières qui appartiennent a ue**.
    18. **on fait un boucle pour faire tout les élèves qui sont dans Gros_Tableau_Notes**.
    19. **la cle est utilisé pour identifier chaque élève dans le tableau Gros_Tableau_Notes**.
    20. **on vérifie si la note de l'élève dans la matière qui correspond bien a une matière**.
    21. **si la cle n'est pas dans notes on met 0**.
    22. **si la cle est dans note on va faire le calcul pour chaque élève**.
    23. **On va utiliser ca pour créer les colonnes des tabeaux** , donc on va suppimer les doublons et on va garder uniquement UE1.1 UE1.2 etc.
    24. **on va faire la même chose pour les étudiants** , on garde un seul et unique étudiant.

    **GÉNÉRATION DU HTML :**

    * **GÉNÉRATION DU HTML** : on prépare la structure HTML.
    * **on parcourt chaque UE dans ues_racines**.
    * **on ajoute <th> sur 4 colonnes pour chaque UE**.
    * **on ajoute 2 colonne pour dire la décission**.
    * **créait 4 colonnes avec les semestre 1 , 2 la moyenne annuelle et état de l'UE**.
    * **on fait une boucle** (sur les étudiants).
    * **on fait un tableau vide pour stocker les moyennes des étudiants**.
    * **on met le Nom et Prénoim dans la première colonne**.
    * **on fait une boucle** (sur les UE).
    * **on construit le nom de la colonne pour UE 1** (et UE 2).
    * **on récupère le nom , prénom pour ue au S1 ou 0.0 si ca existe pas**.
    * **on fait le calcul des moyenne de UE**.
    * **on ajoute cette moyenne annuelle à la liste des moyennes de l'étudiants**.
    * **on détremine le texte d'état de l'UE ( Validé / non validé)**.
    * **On ajoute dans la ligne HTML la note du S1, arrondie à 2 décimales**.
    * **On ajoute dans la ligne HTML la note du S2, arrondie à 2 décimales**.
    * **On ajoute la moyenne annuelle de l'UE, arrondie à 2 décimales**.
    * **on ajoute l'état de l'UE avec la classe css qui correspond à cet état**.
    * **On calcule la décission global de passage en fonction des moyennes annuelles**.
    * **On ajoute la décission de passage dans la dernière colonne**.
    * **On ajoutes toute la ligne HTML de l'étudiant au contenu global**.
    """
    
    #CHEMIN DES FICHIERS EXCEL
    fichier_de_ref = '/workspaces/SAE105_EXCEL_-_VALIDATION_UE/data/coefficients/Coef.xlsx'

    #On créait un tableau vide pour mettre les données qu'on va vouloir utiliser
    dossiers_notes = []
    
    #Chemin pour accéder au note des S1 et S2
    #on met les données qui sont dans le fichier notes_S1 et notes_S2 dans le tableau que l'on vient de créer 
    dossiers_notes.append('/workspaces/SAE105_EXCEL_-_VALIDATION_UE/data/notes/notes_S1')
    dossiers_notes.append('/workspaces/SAE105_EXCEL_-_VALIDATION_UE/data/notes/notes_S2')

    #INITIALISATION DES TABLEAUX
    tableau_coef = []
    note_matiere = []
    Gros_Tableau_Notes = []

    notes = {}

    # Lecture du fichier des coefs sous forme d'un dictionnaire 
    tableau_coef = lecture_du_fichier_coef_dico(fichier_de_ref)

    # Lecture des fichiers des notes des étudiants
    for dossier_notes in dossiers_notes:
        #On va renvoier la liste des fichiers qui sont dans le tableau dossier_notes    
        for fichier in os.listdir(dossier_notes):
            #dans le tableau Gros_Tableau_Notes on rajoute les données de fichier et de dossier_notes
            Gros_Tableau_Notes = Gros_Tableau_Notes + lire_fichier_excel(fichier , dossier_notes)

    # liste des UE
    #on va parcourir tableau_coef et pour chaque case ou y a écrit "Unité_d_Enseignement" on garde la valeur
    liste_ue = list({item["Unité_d_Enseignement"] for item in tableau_coef})
    #On fait la meme chose mais pour "Semestre"
    liste_semestre = list({item["Semestre"] for item in tableau_coef})

    #on fait une boucle pour traiter tout les UE qui sont dans le liste_ue qui est dans le tableau_coef
    for ue in liste_ue:
        #on met en place des filtres
        #on fait une deuxième boucle
        for matière in tableau_coef: 
            matiere_ue = matière["Unité_d_Enseignement"] #on va donner l'UE en fonction de la matière
            matiere_fichier = matière["Fichier"] #on va donner la matière grace a la colonne fichier
            matiere_coef = float(matière["Coefficient"]) # on va dire les coefs de la matière
            if  matiere_ue == ue : # on filtre et on veut que ca soit uniquement les matières qui appartiennent a ue
                for eleve in Gros_Tableau_Notes: # on fait un boucle pour faire tout les élèves qui sont dans Gros_Tableau_Notes
                    cle = (eleve["Nom"], eleve["Prénom"],ue) # la cle est utilisé pour identifier chaque élève dans le tableau Gros_Tableau_Notes
                    if eleve["Fichier_Matière"] == matière["Fichier"] : # on vérifie si la note de l'élève dans la matière qui correspond bien a une matière
                        if cle not in notes: 
                            notes[cle] = 0 # si la cle n'est pas dans notes on met 0
                        notes[cle] += float(eleve["Note"]) * matiere_coef / 100 # si la cle est dans note on va faire le calcul pour chaque élève

    #On va utiliser ca pour créer les colonnes des tabeaux , donc on va suppimer les doublons et on va garder uniquement UE1.1 UE1.2 etc
    ues_racines = sorted(list({ue.split('.')[0] for ue in liste_ue}))
    #on va faire la même chose pour les étudiants , on garde un seul et unique étudiant
    etudiants_uniques = sorted(list(set((k[0], k[1]) for k in notes.keys())))

    # GÉNÉRATION DU HTML 
    html_content = """
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <title>BUT1-R&T</title>
        <style>
            body{ font-family: 'Segoe UI', sans-serif; padding: 20px; font-size: 13px; }
            h2 { text-align: center; color: #333; }
            table{ border-collapse: collapse; width: 100%; box-shadow: 0 0 15px rgba(0,0,0,0.1); margin-top: 20px; }
            td, th{ border: 1px solid #ddd; padding: 8px 4px; text-align: center; }
            thead tr:first-child th { background-color: #005f99; color: white; border-right: 1px solid white; }
            thead tr:nth-child(2) th { background-color: #007acc; color: white; font-size: 0.9em; }
            .col-semestre { background-color: #fcfcfc; color: #555; }
            .col-moyenne { background-color: #fff; font-weight: bold; border-left: 2px solid #ccc; }
            .ue-validee { background-color: #d4edda; color: #155724; font-weight: bold; }
            .ue-compensee { background-color: #fff3cd; color: #856404; font-weight: bold; }
            .ue-echec { background-color: #f8d7da; color: #721c24; font-weight: bold; }
            .decision-ok { background-color: #28a745; color: white; font-weight: bold; font-size: 1.1em; }
            .decision-fail { background-color: #dc3545; color: white; font-weight: bold; font-size: 1.1em; }
            tr:hover { background-color: #f1f1f1; }
        </style>
    </head>
    <body>
        <h2>Notes BUT1 - R&T </h2>
        <table>
        <thead>
            <tr>
                <th rowspan="2" style="width: 150px;">Étudiant</th>
    """

    #on parcourt chaque UE dans ues_racines
    for ue in ues_racines:
        html_content += f'<th colspan="4">{ue} (Annuel)</th>' #on ajoute <th> sur 4 colonnes pour chaque UE

    html_content += '<th rowspan="2">DÉCISION</th></tr>' # on ajoute 2 colonne pour dire la décission
    html_content += '<tr>'
    for ue in ues_racines: 
        html_content += f'<th>{ue}.1</th><th>{ue}.2</th><th>Moy</th><th>État</th>' # créait 4 colonnes avec les semestre 1 , 2 la moyenne annuelle et état de l'UE
    html_content += '</tr></thead><tbody>'


    # on fait une boucle 
    for (nom, prenom) in etudiants_uniques:
        moyennes_annuelles_etudiant = [] # on fait un tableau vide pour stocker les moyennes des étudiants
        ligne_html_etudiant = f"<tr><td style='text-align:left; padding-left:10px;'><b>{nom}</b> {prenom}</td>" # on met le Nom et Prénoim dans la première colonne
        
        #on fait une boucle
        for racine in ues_racines:
            nom_ue_s1 = f"{racine}.1" # on construit le nom de la colonne pour UE 1
            nom_ue_s2 = f"{racine}.2" # on construit le nom de la colonne pour UE 1
            note_s1 = notes.get((nom, prenom, nom_ue_s1), 0.0) #on récupère le nom , prénom pour ue au S1 ou 0.0 si ca existe pas 
            note_s2 = notes.get((nom, prenom, nom_ue_s2), 0.0) #on récupère le nom , prénom pour ue au S2 ou 0.0 si ca existe pas 
            moyenne_annuelle = (note_s1 + note_s2) / 2 # on fait le calcul des moyenne de UE 
            moyennes_annuelles_etudiant.append(moyenne_annuelle) #on ajoute cette moyenne annuelle à la liste des moyennes de l'étudiants
            txt_etat, class_etat = determiner_etat_ue(moyenne_annuelle) # on détremine le texte d'état de l'UE ( Validé / non validé)
            
            ligne_html_etudiant += f'<td class="col-semestre">{round(note_s1, 2)}</td>' # On ajoute dans la ligne HTML la note du S1, arrondie à 2 décimales
            ligne_html_etudiant += f'<td class="col-semestre">{round(note_s2, 2)}</td>' # On ajoute dans la ligne HTML la note du S2, arrondie à 2 décimales
            ligne_html_etudiant += f'<td class="col-moyenne">{round(moyenne_annuelle, 2)}</td>' # On ajoute la moyenne annuelle de l'UE, arrondie à 2 décimales
            ligne_html_etudiant += f'<td class="{class_etat}">{txt_etat}</td>' # on ajoute l'état de l'UE avec la classe css qui correspond à cet état

        txt_decision, class_decision = calculer_decision_passage(moyennes_annuelles_etudiant) # On calcule la décission global de passage en fonction des moyennes annuelles
        ligne_html_etudiant += f'<td class="{class_decision}">{txt_decision}</td></tr>' #On ajoute la décission de passage dans la dernière colonne 
        html_content += ligne_html_etudiant # On ajoutes toute la ligne HTML de l'étudiant au contenu global

    html_content += """
        </tbody>
        </table>
    </body>
    </html>
    """

    with open("/workspaces/SAE105_EXCEL_-_VALIDATION_UE/html/mes_notes.html", "w", encoding="utf-8") as file:
        file.write(html_content)

    print("Le fichier 'mes_notes.html' a été généré avec succès.")

# --- Point d'entrée du script ---
if __name__ == "__main__":
    main()